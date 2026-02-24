#!/usr/bin/env python3
"""
Converte il file Excel dei workshop Didacta in data.json
per la pagina kiosk dello stand C2 Group.

Utilizzo:
  python aggiorna_dati.py percorso_file.xlsx

Requisiti:
  pip install openpyxl
"""

import json
import sys
import os
from datetime import datetime

try:
    from openpyxl import load_workbook
except ImportError:
    print("ERRORE: installa openpyxl con 'pip install openpyxl'")
    sys.exit(1)


def excel_to_json(xlsx_path, output_path="data.json"):
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)

    # Cerca il foglio "AnalisiWorkshop"
    sheet_name = None
    for name in wb.sheetnames:
        if "analisi" in name.lower():
            sheet_name = name
            break

    if not sheet_name:
        # Usa il primo foglio come fallback
        sheet_name = wb.sheetnames[0]
        print(f"⚠ Foglio 'AnalisiWorkshop' non trovato, uso '{sheet_name}'")

    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))

    # Header
    header = [str(h).strip() if h else "" for h in rows[0]]
    print(f"Colonne trovate: {header}")

    # Mappa colonne (flessibile)
    col_map = {}
    aliases = {
        "data": ["data"],
        "time": ["orario", "ora"],
        "type": ["tipologia", "tipo"],
        "code": ["codice", "cod"],
        "partner": ["partner"],
        "room": ["aula", "sala", "room"],
        "title": ["titolo", "title"],
        "abstract": ["abstract", "descrizione"],
        "link": ["link", "url"],
        "capacity": ["capienza", "capacity", "posti"],
        "enrolled": ["iscritti", "enrolled", "registrati"],
    }

    for key, names in aliases.items():
        for i, h in enumerate(header):
            if h.lower() in names:
                col_map[key] = i
                break

    required = ["data", "time", "type", "code", "title"]
    missing = [k for k in required if k not in col_map]
    if missing:
        print(f"ERRORE: colonne mancanti: {missing}")
        print(f"Header: {header}")
        sys.exit(1)

    events = []
    for row in rows[1:]:
        # Salta righe vuote o senza titolo
        code = row[col_map["code"]] if "code" in col_map else None
        if not code:
            continue
        title = row[col_map["title"]] if "title" in col_map else None
        if not title or str(title).strip() == "" or str(title).strip() == "-":
            continue

        # Data
        raw_date = row[col_map["data"]]
        if isinstance(raw_date, datetime):
            date_str = raw_date.strftime("%Y-%m-%d")
        else:
            date_str = str(raw_date).strip()

        # Orario
        raw_time = row[col_map["time"]]
        if isinstance(raw_time, datetime):
            time_str = raw_time.strftime("%H:%M")
        elif hasattr(raw_time, "strftime"):
            time_str = raw_time.strftime("%H:%M")
        else:
            time_str = str(raw_time).strip()

        # Capienza e iscritti
        capacity = 0
        enrolled = 0
        if "capacity" in col_map and row[col_map["capacity"]]:
            try:
                capacity = int(row[col_map["capacity"]])
            except (ValueError, TypeError):
                pass
        if "enrolled" in col_map and row[col_map["enrolled"]]:
            try:
                enrolled = int(row[col_map["enrolled"]])
            except (ValueError, TypeError):
                pass

        event = {
            "date": date_str,
            "time": time_str,
            "type": str(row[col_map["type"]] or "").strip(),
            "code": str(code).strip(),
            "partner": str(row[col_map.get("partner", 0)] or "-").strip(),
            "room": str(row[col_map.get("room", 0)] or "").strip(),
            "title": str(row[col_map["title"]] or "").strip(),
            "abstract": str(row[col_map.get("abstract", 0)] or "").strip(),
            "link": str(row[col_map.get("link", 0)] or "").strip(),
            "capacity": capacity,
            "enrolled": enrolled,
        }
        events.append(event)

    # Ordina per data e orario
    events.sort(key=lambda e: (e["date"], e["time"], e["code"]))

    with open(output_path, "w", encoding="utf-8") as f:
        json.dump(events, f, ensure_ascii=False, indent=2)

    print(f"\n✅ Generato {output_path} con {len(events)} eventi")

    # Report
    dates = sorted(set(e["date"] for e in events))
    for d in dates:
        day_events = [e for e in events if e["date"] == d]
        sold_out = sum(1 for e in day_events if e["capacity"] > 0 and e["enrolled"] >= e["capacity"])
        print(f"   {d}: {len(day_events)} eventi ({sold_out} sold out)")

    return events


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Utilizzo: python aggiorna_dati.py <file.xlsx>")
        sys.exit(1)

    xlsx = sys.argv[1]
    if not os.path.exists(xlsx):
        print(f"ERRORE: file non trovato: {xlsx}")
        sys.exit(1)

    # Output nella stessa directory dello script
    script_dir = os.path.dirname(os.path.abspath(__file__))
    output = os.path.join(script_dir, "data.json")

    excel_to_json(xlsx, output)
